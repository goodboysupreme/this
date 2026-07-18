#!/usr/bin/env python
"""Load the real BITS Pilani datasets into the PlaceIQ database.

Sources (all under ``data/raw/``):

- ``ps2_24-25_stats.csv``           — PS-2 allotments 2024-25 (ps2.pilani.online
                                      export): branch, CGPA, station, stipend,
                                      semester flags.
- ``SI_Tracker_Official.xlsx``      — official SI tracker, sheets 25-26 / 24-25 /
                                      23-24: company, branch, monthly CTC, role,
                                      notes, CG & branch cutoff text.
- ``22-23_Both_Sems.pdf``           — Placement Unit "Internship Chronicles"
- ``23-24_Sem1_SI_Chronicles.pdf``  — (per-company interview experiences)
- ``23-24_Sem2_Placements_SI_Chronicles.pdf``
- ``24-25.pdf``

PRIVACY: the raw sources contain real student names. Student names are NEVER
stored in the database — the tracker ``Name`` column is dropped on read, and
experience content has name-like lines stripped on a best-effort basis (lines
starting with ``Interviewee``/``Name:``, BITS-ID tokens, and bare person-name
lines). ``author_hint`` is always a non-identifying label such as
"Placement Unit Chronicles 24-25".

Mock coexistence: this script OWNS the companies / offers / cutoff_stats /
experiences tables. Every run wipes those four tables completely (this also
removes the deterministic mock seed from ``scripts/seed_mock.py``) and reloads
them from the real sources. Users, favorites, contacts, email templates,
campaigns and page_events are preserved. The test-suite DB
(``tests/conftest.py``) seeds mock data into its own throwaway SQLite file and
is unaffected.

Usage:
    cd backend && python ../scripts/load_real_data.py [DATABASE_URL]

Default DATABASE_URL: sqlite at ``backend/placeiq.db``.
Idempotent: safe to re-run; owned tables are wiped and rebuilt each run.
"""

from __future__ import annotations

import csv
import re
import sys
from difflib import get_close_matches
from pathlib import Path
from statistics import median

BACKEND_DIR = Path(__file__).resolve().parent.parent / "backend"
DATA_RAW = Path(__file__).resolve().parent.parent / "data" / "raw"
sys.path.insert(0, str(BACKEND_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import fitz  # noqa: E402  (PyMuPDF)
import openpyxl  # noqa: E402
from sqlalchemy import create_engine, delete, select  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402

from app.db import Base  # noqa: E402
from app.models import Company, CutoffStat, Experience, Offer  # noqa: E402
from seed_mock import slugify  # noqa: E402  keep slugs consistent with app code

# ---------------------------------------------------------------------------
# Company name normalization
# ---------------------------------------------------------------------------

# Trailing segments after a comma/dash that are locations or team descriptors
# are stripped: "Nutanix , Bengaluru / Pune" -> "Nutanix",
# "Goldman Sachs - Hyderabad" -> "Goldman Sachs".

# Canonical-name aliases, keyed by "merge key" (lowercase, alnum only).
# Merges obvious dupes across sources ("D.E. Shaw" / "DE Shaw", "PnG" /
# "Procter & Gamble", tracker abbreviations, etc.).
CANONICAL: dict[str, str] = {}


def _register(canonical: str, *keys: str) -> None:
    for key in keys:
        CANONICAL[key] = canonical


_register("D.E. Shaw", "deshaw")
_register("JPMorgan Chase", "jpmc", "jpmorgan", "jpmorganchase", "jpmorganchaseco")
_register("Procter & Gamble", "png", "proctergamble", "procterandgamble", "pg")
_register("Boston Consulting Group (BCG)", "bcg", "bostonconsultinggroup")
_register("McKinsey & Company", "mckinsey", "mckinseycompany", "mckinseyandcompany")
_register("Hindustan Unilever", "hul", "hindustanunilever", "hindustanunileverlimited")
_register("Amazon", "amazon", "amazonsde")
_register("Google", "google", "googlestep", "googlestepppi")
_register("Microsoft", "microsoft")
_register("Nvidia", "nvidia", "nvidiagraphics", "nvidiahardware")
_register("Texas Instruments", "texasinstruments")
_register("Qualcomm", "qualcomm")
_register(
    "Samsung R&D",
    "samsung", "samsungrd", "samsungrdinstitute", "samsungrdindia",
    "samsungresearch", "samsungsemiconductor",
)
_register(
    "Accenture", "accenture", "accenturesc", "accenturestrategyconsulting",
    "accenturestrategyandconsulting",
)
_register("Mondelez International", "mondelez", "mondelezinternational", "mondelzinternational")
_register("UBS", "ubs")
_register("Zomato", "zomato", "zomatogrofers")
_register("Flipkart", "flipkart")
_register("Walmart Global Tech", "walmart", "walmartglobaltech")
_register("Goldman Sachs", "goldmansachs")
_register("Nomura", "nomura")
_register("Oracle", "oracle")
_register("Micron Technology", "micron", "microntechnology")
_register("Tower Research Capital", "towerresearch", "towerresearchcapital")
_register("HDFC Bank", "hdfc", "hdfcbank")
_register("LTIMindtree", "ltimindtree", "larsentoubroinfotech")
_register("Schlumberger (SLB)", "schlumberger", "slb")
_register("Société Générale", "socitgnrale", "societegenerale", "socgen")
_register("Jaguar Land Rover", "jaguarlandrover")
_register("Myntra", "myntra", "myntradesigns")
_register("TCS", "tcs", "tataconsultancyservices")
_register("Adobe", "adobe")
_register("Atlassian", "atlassian")
_register("Intuit", "intuit")
_register("Uber", "uber")
_register("Cisco", "cisco")
_register("Salesforce", "salesforce")
_register("Rubrik", "rubrik")
_register("Sprinklr", "sprinklr")
_register("Arcesium", "arcesium")
_register("Capital One", "capitalone")
_register("Wells Fargo", "wellsfargo")
_register("Tata Steel", "tatasteel")
_register("ITC", "itc")
_register("JSW Steel", "jsw", "jswsteel")
_register("ExxonMobil", "exxonmobil")
_register("Shell", "shell")
_register("Expedia Group", "expedia", "expediagroup")
_register("InfoEdge", "infoedge", "infoedgeindia")
_register("DMI Finance", "dmifinance")
_register("LimeChat", "limechat")
_register("Deloitte", "deloitte")
_register("KPMG", "kpmg")
_register("Apple", "apple")
_register("Intel", "intel")
_register("Nutanix", "nutanix")
_register("MediaTek", "mediatek")
_register("Luminous Power Technologies", "luminous", "luminouspowertechnologies")

# Alias keys long enough for safe prefix matching ("jpmccibra" -> JPMorgan).
_PREFIX_KEYS = sorted((k for k in CANONICAL if len(k) >= 4), key=len, reverse=True)

_LEGAL_SUFFIX_RE = re.compile(
    r"[\s,]*(\((pvt\.?|private)\s*ltd\.?\)|pvt\.?\s*ltd\.?|private\s+limited|limited|ltd\.?|inc\.?|llp|llc)\s*$",
    re.IGNORECASE,
)


def _merge_key(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


def _smart_case(name: str) -> str:
    """Title-case ALL-CAPS names, keeping short tokens (CSIR, R&D) upper."""
    if not name.isupper():
        return name

    def fix_word(word: str) -> str:
        parts = word.split("-")
        return "-".join(p if len(p) <= 4 else p.title() for p in parts)

    return " ".join(fix_word(w) for w in name.split())


def normalize_company(raw: str) -> str:
    """Normalize a raw station/company string to a canonical company name."""
    name = str(raw).replace("�", "-").strip()
    name = re.sub(r"\([^)]*\)", "", name)  # strip parentheticals
    name = name.split(",")[0]  # drop location tail after first comma
    # Drop " - <location/descriptor>" tails ("TI - Analog", "GS - Hyderabad").
    name = re.split(r"\s[-–—]+\s*|\s*[-–—]+\s", name)[0]
    name = _LEGAL_SUFFIX_RE.sub("", name).strip(" -–—,")
    name = re.sub(r"\s+", " ", name).strip()
    key = _merge_key(name)
    if key in CANONICAL:
        return CANONICAL[key]
    for prefix in _PREFIX_KEYS:
        if key.startswith(prefix):
            return CANONICAL[prefix]
    return _smart_case(name)


# ---------------------------------------------------------------------------
# Sector & role-category inference
# ---------------------------------------------------------------------------

_SECTOR_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("Finance", (
        "goldman", "nomura", "jpmorgan", "jpmc", "morgan stanley", "wells fargo",
        "d.e. shaw", "tower research", "citadel", "hsbc", "barclays", "deutsche",
        "ubs", "credit suisse", "american express", "mastercard", "visa", "hdfc",
        "icici", "axis", "kotak", "société", "socgen", "dmi finance", "capital one",
        "fidelity", "blackrock", "jane street", "optiver", "alphagrep", "graviton",
        "quadeye", "arcesium", "nuvama", "edelweiss", "avr capital", "plutus",
    )),
    ("Semiconductors", (
        "nvidia", "qualcomm", "texas instruments", "micron", "intel", "amd",
        "broadcom", "mediatek", "arm", "cadence", "synopsys", "analog devices",
        "nxp", "infineon", "samsung",
    )),
    ("Consulting", (
        "mckinsey", "bcg", "boston consulting", "bain", "deloitte", "kpmg",
        "ernst", "pwc", "accenture", "praxis", "kearney", "zs associates",
        "nation with namo",
    )),
    ("FMCG", (
        "hindustan unilever", "procter", "mondelez", "itc", "nestle", "nestlé",
        "dabur", "marico", "colgate", "pepsico", "coca", "unilever",
    )),
    ("Tech", (
        "google", "microsoft", "amazon", "adobe", "oracle", "cisco", "salesforce",
        "atlassian", "intuit", "uber", "flipkart", "walmart", "zomato", "myntra",
        "sprinklr", "rubrik", "apple", "meta", "linkedin", "ola", "swiggy",
        "phonepe", "paytm", "nutanix", "servicenow", "vmware", "limechat",
        "infoedge", "tcs", "ltimindtree", "infosys", "wipro", "capgemini",
        "cognizant", "hcl", "adobe", "postman", "razorpay", "groww", "cred",
        "dunzo", "freshworks", "zeta", "tekion", " Rippling", "rippling",
    )),
    ("Analytics", (
        "fractal", "tiger analytics", "mu sigma", "latentview", "affine",
        "bridgei2i", "course5",
    )),
    ("Core", (
        "tata steel", "jsw", "shell", "exxonmobil", "schlumberger", "siemens",
        "bosch", "jaguar land rover", "mercedes", "tata motors", "mahindra",
        "reliance", "aditya birla", "hindalco", "vedanta", "csir", "drdo",
        "isro", "luminous", "larsen", "sterlite", "hindustan zinc", "hero",
        "bajaj", "tvs", "ashok leyland", "gail", "ongc", "bpcl", "hpcl",
        "exxon", "slb", " Eaton", "eaton", "honeywell", "ge ", "abb",
    )),
    ("Electronics", (
        "luminous", "bosch", "havells", "philips", "lg ", "panasonic",
    )),
]


def infer_sector(company_name: str) -> str:
    low = company_name.lower()
    for sector, keywords in _SECTOR_RULES:
        if any(k.strip() and k in low for k in keywords):
            return sector
    return "other"


_ROLE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("finance", (
        "quant", "finance", "trading", "bank", "investment", "risk", "equity",
        "capital market", "treasury", "actuar", "portfolio",
    )),
    ("consulting", ("consult",)),
    ("analytics", (
        "analy", "data scien", "machine learning", "data eng", "business intel",
        "research intern", "data analyst", "ai ", " ml ",
    )),
    ("sde", (
        "sde", "software", "swe", "develop", "backend", "frontend", "full stack",
        "fullstack", "devops", "cloud", "platform eng", "technology",
    )),
    ("core", (
        "analog", "digital", "vlsi", "embedded", "hardware", "rtl",
        "verification", "design eng", "thermal", "structural", "manufacturing",
        "supply chain", "process eng", "mechanical", "civil", "chemical",
        "electrical", "power", "signal",
    )),
]

_SECTOR_TO_ROLE = {
    "Finance": "finance",
    "Consulting": "consulting",
    "Analytics": "analytics",
    "Tech": "sde",
    "Semiconductors": "core",
    "Electronics": "core",
    "Core": "core",
}


def infer_role_category(text: str, sector: str | None = None) -> str:
    low = f" {text.lower()} "
    for category, keywords in _ROLE_RULES:
        if any(k in low for k in keywords):
            return category
    return _SECTOR_TO_ROLE.get(sector or "", "other")


# ---------------------------------------------------------------------------
# Branch mapping (BITS codes)
# ---------------------------------------------------------------------------

BRANCH_CODES = {
    "A7": "CS",
    "A1": "Chemical",
    "A2": "Civil",
    "A3": "EEE",
    "A4": "Mechanical",
    "A8": "E&I",
    "AA": "ECE",
    "B1": "Biology",
    "B2": "Chemistry",
    "B3": "Economics",
    "B4": "Math",
    "B5": "Physics",
}


def map_branch(code: object) -> str | None:
    """Map a BITS branch code to a readable branch; keep raw code if unknown.

    Composite dual-degree codes (``B1A1``) map each 2-char part and join.
    """
    if code is None:
        return None
    text = str(code).strip().upper()
    if not text or text == "NAN":
        return None
    if text in BRANCH_CODES:
        return BRANCH_CODES[text]
    if len(text) % 2 == 0:
        parts = [text[i : i + 2] for i in range(0, len(text), 2)]
        if all(part in BRANCH_CODES for part in parts):
            return " + ".join(BRANCH_CODES[part] for part in parts)
    return text  # unknown code (A3TS, D2, ...) — keep raw


# ---------------------------------------------------------------------------
# Company registry
# ---------------------------------------------------------------------------


class CompanyRegistry:
    """Collects canonical companies; assigns unique slugs on insert."""

    def __init__(self) -> None:
        self.by_key: dict[str, dict] = {}

    def get_or_create(self, raw_name: str, sector_hint: str | None = None) -> dict:
        name = normalize_company(raw_name)
        key = _merge_key(name)
        entry = self.by_key.get(key)
        if entry is None:
            sector = infer_sector(name)
            if sector == "other" and sector_hint:
                sector = sector_hint
            entry = {"name": name, "sector": sector}
            self.by_key[key] = entry
        elif entry["sector"] == "other" and sector_hint:
            entry["sector"] = sector_hint
        return entry

    def find(self, raw_name: str) -> dict | None:
        """Exact then fuzzy lookup of a company in the registry."""
        key = _merge_key(normalize_company(raw_name))
        if key in self.by_key:
            return self.by_key[key]
        close = get_close_matches(key, self.by_key.keys(), n=1, cutoff=0.9)
        return self.by_key[close[0]] if close else None


# ---------------------------------------------------------------------------
# Source 1: PS-2 CSV
# ---------------------------------------------------------------------------


def load_ps2_offers(registry: CompanyRegistry) -> tuple[list[dict], dict]:
    path = DATA_RAW / "ps2_24-25_stats.csv"
    offers: list[dict] = []
    stats = {"rows": 0, "skipped": 0}
    with path.open(newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            stats["rows"] += 1
            station = (row.get("AllottedStationName") or "").strip()
            cgpa = _to_float(row.get("CGPA"))
            if not station or cgpa is None:
                stats["skipped"] += 1
                continue
            entry = registry.get_or_create(station)
            stipend = _to_float(row.get("Stipend"))
            sem1 = _to_float(row.get("AllotedSemester1")) == 1.0
            offers.append(
                {
                    "company_key": _merge_key(entry["name"]),
                    "type": "ps2",
                    # Session 2024-25: semester 1 -> 2024, semester 2 -> 2025.
                    "year": 2024 if sem1 else 2025,
                    "role": "PS-II Station",
                    "role_category": infer_role_category(station, entry["sector"]),
                    "branch": map_branch(row.get("Branch")),
                    # This row's allotment CG; the per-station MIN lands in
                    # cutoff_stats.
                    "cgpa_cutoff": cgpa,
                    "stipend_ctc": round(stipend / 1000.0, 1) if stipend else None,
                    "slots": None,
                }
            )
    return offers, stats


# ---------------------------------------------------------------------------
# Source 2: SI tracker XLSX
# ---------------------------------------------------------------------------

_SHEET_YEARS = {"25-26": 2025, "24-25": 2024, "23-24": 2023}
_JUNK_COMPANIES = {
    "total", "average", "median", "stats", "statistics", "company", "na", "n/a",
}


def _header_map(header_row: tuple) -> dict[str, int]:
    """Map canonical fields to column indexes for one SI tracker sheet."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        header = str(raw).strip().lower() if raw is not None else ""
        if not header:
            continue
        if header == "company":
            mapping.setdefault("company", idx)
        elif header == "name":
            mapping.setdefault("name", idx)  # dropped — privacy
        elif header == "branch":
            mapping.setdefault("branch", idx)
        elif "ctc" in header or "stipend" in header:
            mapping.setdefault("stipend", idx)
        elif header == "role":
            mapping.setdefault("role", idx)
        elif header in ("notes", "remarks"):
            mapping.setdefault("notes", idx)
        elif "cg & branch cutoff" in header:
            mapping.setdefault("cgcut", idx)
    return mapping


def load_si_offers(registry: CompanyRegistry) -> tuple[list[dict], dict]:
    path = DATA_RAW / "SI_Tracker_Official.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    offers: list[dict] = []
    stats: dict[str, int] = {"rows": 0, "skipped_no_stipend": 0, "skipped_other": 0}
    per_sheet: dict[str, int] = {}

    for sheet, year in _SHEET_YEARS.items():
        if sheet not in wb.sheetnames:
            print(f"  ! sheet '{sheet}' missing in SI tracker — skipped")
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            continue
        cols = _header_map(header)
        if "company" not in cols or "stipend" not in cols:
            print(f"  ! sheet '{sheet}': unexpected headers {header!r} — skipped")
            continue

        company_cell: str | None = None
        kept = 0
        for row in rows:
            stats["rows"] += 1
            raw_company = row[cols["company"]] if cols["company"] < len(row) else None
            if raw_company is not None and str(raw_company).strip():
                company_cell = str(raw_company).strip()
            if not company_cell or company_cell.strip().lower() in _JUNK_COMPANIES:
                continue

            stipend_raw = row[cols["stipend"]] if cols["stipend"] < len(row) else None
            stipend = _to_float(stipend_raw)
            if stipend is None:
                # Non-numeric / missing stipend rows (headers, totals, blank
                # trailing rows) — counted, not stored.
                stats["skipped_no_stipend"] += 1
                continue

            def cell(field: str) -> str | None:
                idx = cols.get(field)
                if idx is None or idx >= len(row) or row[idx] is None:
                    return None
                text = str(row[idx]).strip()
                return text or None

            # Name column is never read into the offer — privacy.
            notes = cell("notes")
            cgcut_text = cell("cgcut")
            if cgcut_text:
                notes = f"{notes}; CG & branch cutoff: {cgcut_text}" if notes else (
                    f"CG & branch cutoff: {cgcut_text}"
                )
            cgpa = None
            if cgcut_text:
                match = re.search(r"(\d+(?:\.\d+)?)\s*\+", cgcut_text)
                if match:
                    cgpa = float(match.group(1))

            entry = registry.get_or_create(company_cell)
            role = cell("role")
            offers.append(
                {
                    "company_key": _merge_key(entry["name"]),
                    "type": "si",
                    "year": year,
                    "role": role or "Summer Intern",
                    "role_category": infer_role_category(
                        f"{role or ''} {company_cell}", entry["sector"]
                    ),
                    "branch": cell("branch"),
                    "cgpa_cutoff": cgpa,
                    "stipend_ctc": round(stipend / 1000.0, 1),
                    "slots": None,
                    "notes": notes,
                }
            )
            kept += 1
        per_sheet[sheet] = kept
    stats["per_sheet"] = per_sheet  # type: ignore[assignment]
    return offers, stats


# ---------------------------------------------------------------------------
# Cutoff stats (same aggregation as scripts/seed_mock.py)
# ---------------------------------------------------------------------------


def build_cutoff_stats(offers: list[dict]) -> list[dict]:
    groups: dict[tuple[str, str, int], list[float]] = {}
    for offer in offers:
        if offer.get("cgpa_cutoff") is None:
            continue
        key = (offer["company_key"], offer["type"], offer["year"])
        groups.setdefault(key, []).append(offer["cgpa_cutoff"])

    stats: list[dict] = []
    for (company_key, offer_type, year), cgs in sorted(groups.items()):
        cgs = sorted(cgs)
        stats.append(
            {
                "company_key": company_key,
                "type": offer_type,
                "year": year,
                "min_cg": round(cgs[0], 2),
                "median_cg": round(median(cgs), 2),
                "p25": round(cgs[0] + (cgs[-1] - cgs[0]) * 0.25, 2),
                "p75": round(cgs[0] + (cgs[-1] - cgs[0]) * 0.75, 2),
            }
        )
    return stats


# ---------------------------------------------------------------------------
# Source 3: Placement Unit Chronicles PDFs -> experiences
# ---------------------------------------------------------------------------

# Files using the "running company-name page header" template.
# (filename, type, year, author_hint, category headers)
_HEADER_TEMPLATE_FILES = [
    (
        "24-25.pdf", "si", 2024, "Placement Unit Chronicles 24-25",
        {
            "consulting", "core", "data science", "electronics technology",
            "general management", "information technology",
            "quantitative finance", "supply chain management",
        },
    ),
    (
        "23-24_Sem1_SI_Chronicles.pdf", "si", 2023,
        "Placement Unit Chronicles 23-24 Sem 1",
        {
            "analytics", "consulting", "core", "data science", "electronics",
            "it", "manufacturing and scm", "manufacturing", "quant",
        },
    ),
    # Semester-2 23-24 chronicles is explicitly On-Campus PLACEMENTS
    # (its own disclaimer: "contains entries only for On-Campus
    # Placements"), so its sections become type="placement".
    (
        "23-24_Sem2_Placements_SI_Chronicles.pdf", "placement", 2023,
        "Placement Unit Chronicles 23-24 Sem 2",
        {
            "analytics", "consulting", "core", "electronics technology",
            "finance and quant", "information technology",
            "supply chain and general management", "product management",
            "placements",
        },
    ),
]

_FRONT_MATTER = {
    "table of contents", "chronicles", "meet the team", "disclaimer",
    "summer internships", "placements", "internship chronicles",
}

_HEADING_WORDS = {
    "introduction", "personal experience", "selection process",
    "eligibility criteria", "recruitment process", "recruitment procedure",
    "job role", "job description", "job profile", "disclaimer",
    "table of contents", "words of advice", "sources of preparation",
    "relevant courses", "company description", "interview process",
    "online assessment", "technical interview", "hr interview",
    "group discussion", "case study", "present and previous recruiters",
    "meet the team", "resources", "summary", "overview", "conclusion",
}

_PERSON_NAME_RE = re.compile(r"[A-Z][a-zA-Z.'-]+( [A-Z][a-zA-Z.'-]+){1,2}")
_BITS_ID_RE = re.compile(r"\(?\b20\d{2}[A-Z0-9]{6,10}P\b\)?")


def _collapse(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().lower()


def _is_person_name_line(line: str) -> bool:
    """Bare person-name heuristic: 2-3 title-cased words, no digits, not a
    known section heading, short. Best-effort privacy filter."""
    if len(line) > 40 or _collapse(line) in _HEADING_WORDS:
        return False
    return bool(_PERSON_NAME_RE.fullmatch(line))


def _clean_experience_lines(lines: list[str]) -> list[str]:
    out: list[str] = []
    for line in lines:
        text = line.strip()
        if not text or text.isdigit():
            continue
        # Drop interviewee/name attribution lines entirely (privacy).
        if re.match(r"^(interviewee|name)\s*[-:]", text, re.IGNORECASE):
            continue
        # Strip BITS-ID tokens anywhere ("(2020A4PS0575P)").
        if _BITS_ID_RE.search(text):
            text = _BITS_ID_RE.sub("", text).strip(" -–—()")
            if not text:
                continue
        if _is_person_name_line(text):
            continue
        out.append(text)
    return out


def _finalize_content(lines: list[str]) -> str | None:
    content = "\n".join(_clean_experience_lines(lines))
    content = re.sub(r"\n{3,}", "\n\n", content).strip()
    if len(content) < 20:
        return None
    if len(content) > 4900:
        cut = content[:4900]
        content = cut[: cut.rfind("\n")] if "\n" in cut else cut
    return content


def _category_sector(category: str | None) -> str | None:
    if not category:
        return None
    cat = category.lower()
    if "core" in cat or "supply" in cat or "manufactur" in cat:
        return "Core"
    if "electronic" in cat:
        return "Electronics"
    if cat in ("it",) or "information" in cat:
        return "Tech"
    if "analy" in cat or "data" in cat:
        return "Analytics"
    if "consult" in cat:
        return "Consulting"
    if "quant" in cat or "finance" in cat:
        return "Finance"
    if "product" in cat:
        return "Tech"
    return None


def _page_lines(page) -> list[str]:
    return [l.strip() for l in page.get_text().splitlines() if l.strip()]


def parse_header_template_pdf(
    path: Path, categories: set[str]
) -> tuple[list[dict], list[str]]:
    """Parse chronicles where each body page's first line is the company name.

    Returns (sections, warnings). Each section: {company, category, lines}.
    """
    warnings: list[str] = []
    doc = fitz.open(path)
    sections: list[dict] = []
    current_company: str | None = None
    current_category: str | None = None
    buf: list[str] = []
    seen_toc = False

    def flush() -> None:
        nonlocal buf
        if current_company and buf:
            sections.append(
                {"company": current_company, "category": current_category, "lines": buf}
            )
        buf = []

    for page in doc:
        lines = _page_lines(page)
        if not lines:
            continue
        # Locate the running header: first line, unless it's a page number.
        header_idx = 1 if re.fullmatch(r"\d{1,3}", lines[0]) and len(lines) > 1 else 0
        header = lines[header_idx]
        body = lines[header_idx + 1 :]
        if body and re.fullmatch(r"\d{1,3}", body[0]):
            body = body[1:]

        header_low = _collapse(header)
        joined = " ".join(_collapse(l) for l in lines)

        if not seen_toc:
            if "table of contents" in joined:
                seen_toc = True
            continue

        is_category_page = "present and previous recruiters" in joined
        if header_low in categories or is_category_page:
            flush()
            current_company = None
            if header_low in categories:
                current_category = header_low
            continue
        if header_low in _FRONT_MATTER:
            flush()
            current_company = None
            continue

        if header != current_company:
            flush()
            current_company = header
        buf.extend(body)
    flush()
    doc.close()

    if not sections:
        warnings.append(f"{path.name}: no company sections found — file skipped")
    return sections, warnings


def parse_22_23_pdf(path: Path) -> tuple[list[dict], list[str]]:
    """Parse the 22-23 chronicles ("Company: X" section markers template)."""
    warnings: list[str] = []
    doc = fitz.open(path)
    sections: list[dict] = []
    current_category: str | None = None
    current_company: str | None = None
    buf: list[str] = []
    started = False

    def flush() -> None:
        nonlocal buf
        if current_company and buf:
            sections.append(
                {"company": current_company, "category": current_category, "lines": buf}
            )
        buf = []

    for page in doc:
        for line in _page_lines(page):
            low = _collapse(line)
            sector_match = re.match(r"^sector\s*:\s*(.+)$", line, re.IGNORECASE)
            company_match = re.match(r"^company\s*:\s*(.+)$", line, re.IGNORECASE)
            if sector_match:
                current_category = _collapse(sector_match.group(1))
                continue
            if company_match:
                flush()
                started = True
                current_company = company_match.group(1).strip()
                continue
            if not started:
                continue  # cover + TOC pages
            # Skip all-caps divider page text ("ANALYTICS").
            if line.isupper() and len(line) < 40:
                continue
            buf.append(line)
    flush()
    doc.close()

    if not sections:
        warnings.append(f"{path.name}: no 'Company:' sections found — file skipped")
    return sections, warnings


def _split_interviewees(lines: list[str]) -> list[list[str]]:
    """Split a company section into per-interviewee chunks.

    Chronicles repeat Introduction / Interviewee / Selection Process blocks
    per student; each block becomes one Experience row.
    """
    chunks: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        if re.match(r"^interviewee\s*[-:]", line, re.IGNORECASE) and current:
            chunks.append(current)
            current = []
        current.append(line)
    if current:
        chunks.append(current)
    return chunks or [lines]


def load_experiences(
    registry: CompanyRegistry,
) -> tuple[list[dict], dict[str, int], list[str]]:
    experiences: list[dict] = []
    per_file: dict[str, int] = {}
    warnings: list[str] = []
    created_companies = 0

    def emit(sections: list[dict], exp_type: str, year: int, author: str, fname: str) -> None:
        nonlocal created_companies
        count = 0
        for section in sections:
            company_name = section["company"]
            sector_hint = _category_sector(section.get("category"))
            entry = registry.find(company_name)
            if entry is None:
                # Strong TOC/section company missing from offer sources —
                # create it so the experience has a home.
                entry = registry.get_or_create(company_name, sector_hint=sector_hint)
                created_companies += 1
            for chunk in _split_interviewees(section["lines"]):
                content = _finalize_content(chunk)
                if content is None:
                    continue
                experiences.append(
                    {
                        "company_key": _merge_key(entry["name"]),
                        "type": exp_type,
                        "year": year,
                        "author_hint": author,
                        "content": content,
                        "approved": True,
                    }
                )
                count += 1
        per_file[fname] = count

    # 22-23: "Company:" marker template (type=si — Internship Chronicles).
    sections, warns = parse_22_23_pdf(DATA_RAW / "22-23_Both_Sems.pdf")
    warnings.extend(warns)
    if sections:
        emit(sections, "si", 2022, "Placement Unit Chronicles 22-23", "22-23_Both_Sems.pdf")

    for fname, exp_type, year, author, categories in _HEADER_TEMPLATE_FILES:
        path = DATA_RAW / fname
        if not path.exists():
            warnings.append(f"{fname}: file missing — skipped")
            continue
        sections, warns = parse_header_template_pdf(path, categories)
        warnings.extend(warns)
        if sections:
            emit(sections, exp_type, year, author, fname)

    return experiences, per_file, warnings


# ---------------------------------------------------------------------------
# Small numeric helpers
# ---------------------------------------------------------------------------


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("₹", "").replace(",", "")
    if not text or text.lower() in ("na", "n/a", "-", "--", "nan", "none", "unpaid"):
        return None
    try:
        return float(text)
    except ValueError:
        match = re.search(r"\d+(?:\.\d+)?", text)
        return float(match.group()) if match else None


# ---------------------------------------------------------------------------
# Main load
# ---------------------------------------------------------------------------


def wipe_owned_tables(session) -> dict[str, int]:
    """Delete ALL companies/offers/cutoff_stats/experiences (mock or real).

    This script owns those four tables. Users, favorites, contacts,
    templates, campaigns and page_events are deliberately preserved.
    """
    counts = {}
    for model in (Experience, CutoffStat, Offer, Company):
        counts[model.__tablename__] = session.query(model).count()
        session.execute(delete(model))
    session.commit()
    return counts


def load(session) -> dict:
    summary: dict = {}

    wiped = wipe_owned_tables(session)
    summary["wiped"] = wiped

    registry = CompanyRegistry()

    ps2_offers, ps2_stats = load_ps2_offers(registry)
    si_offers, si_stats = load_si_offers(registry)
    offers = ps2_offers + si_offers
    cutoff_stats = build_cutoff_stats(offers)
    experiences, per_file_exp, exp_warnings = load_experiences(registry)

    # Insert companies with unique slugs.
    id_by_key: dict[str, int] = {}
    used_slugs: set[str] = set()
    for key in sorted(registry.by_key, key=lambda k: registry.by_key[k]["name"]):
        data = registry.by_key[key]
        slug = slugify(data["name"])
        base, n = slug, 2
        while slug in used_slugs:
            slug = f"{base}-{n}"
            n += 1
        used_slugs.add(slug)
        company = Company(
            name=data["name"],
            slug=slug,
            sector=data["sector"],
            description=(
                f"{data['name']} — {data['sector']} recruiter at BITS Pilani "
                "(real data)."
            ),
        )
        session.add(company)
        session.flush()
        id_by_key[key] = company.id

    for data in offers:
        row = {k: v for k, v in data.items() if k != "notes"}  # Offer has no notes col
        row["company_id"] = id_by_key[row.pop("company_key")]
        session.add(Offer(**row))

    for data in cutoff_stats:
        row = dict(data)
        row["company_id"] = id_by_key[row.pop("company_key")]
        session.add(CutoffStat(**row))

    for data in experiences:
        row = dict(data)
        row["company_id"] = id_by_key[row.pop("company_key")]
        session.add(Experience(**row))

    session.commit()

    summary["ps2"] = ps2_stats
    summary["si"] = si_stats
    summary["companies"] = len(registry.by_key)
    summary["offers_by_type"] = {
        t: sum(1 for o in offers if o["type"] == t) for t in ("ps2", "si")
    }
    summary["cutoff_stats"] = len(cutoff_stats)
    summary["experiences_per_file"] = per_file_exp
    summary["experience_warnings"] = exp_warnings
    return summary


def print_summary(summary: dict, session) -> None:
    print("\n=== Real data load summary ===")
    print(f"Wiped previous rows: {summary['wiped']}")
    ps2, si = summary["ps2"], summary["si"]
    print(f"\nPS-2 CSV (ps2_24-25_stats.csv): {ps2['rows']} rows read, "
          f"{ps2['rows'] - ps2['skipped']} offers loaded, {ps2['skipped']} skipped")
    print(f"SI tracker (SI_Tracker_Official.xlsx): {si['rows']} rows read, "
          f"{si['skipped_no_stipend']} skipped (non-numeric/blank stipend), "
          f"per sheet kept: {si['per_sheet']}")
    print(f"\nCompanies: {summary['companies']}")
    print(f"Offers by type: {summary['offers_by_type']}")
    print(f"Cutoff stats rows: {summary['cutoff_stats']}")
    print(f"Experiences per file: {summary['experiences_per_file']}")
    for warning in summary["experience_warnings"]:
        print(f"  WARNING: {warning}")

    print("\n--- Sample rows ---")
    for company in session.query(Company).order_by(Company.name).limit(8):
        print(f"  company: {company.name!r} slug={company.slug!r} sector={company.sector!r}")
    for offer in (
        session.query(Offer).filter(Offer.type == "ps2").limit(3).all()
        + session.query(Offer).filter(Offer.type == "si").limit(3).all()
    ):
        print(
            f"  offer: {offer.type} {offer.year} company_id={offer.company_id} "
            f"role={offer.role!r} branch={offer.branch!r} cg={offer.cgpa_cutoff} "
            f"pay={offer.stipend_ctc}k cat={offer.role_category!r}"
        )
    exp = session.query(Experience).first()
    if exp:
        snippet = exp.content[:160].replace("\n", " | ")
        print(f"  experience: {exp.type} {exp.year} hint={exp.author_hint!r}\n    {snippet}...")


def main() -> None:
    db_url = (
        sys.argv[1]
        if len(sys.argv) > 1
        else f"sqlite:///{BACKEND_DIR.as_posix()}/placeiq.db"
    )
    engine = create_engine(db_url, connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine)()
    try:
        summary = load(session)
        print_summary(summary, session)
    finally:
        session.close()


if __name__ == "__main__":
    main()
